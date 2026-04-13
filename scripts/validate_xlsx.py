import glob
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is not installed. Please install it using 'pip install openpyxl'")
    sys.exit(1)

def validate_excel_file(filepath):
    print(f"Validating {filepath}...")
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        # Check if the workbook has sheets
        if not wb.sheetnames:
            print(f"  [ERROR] {filepath} has no sheets.")
            return False
            
        # Get the active sheet
        ws = wb.active
        
        # Check for data in cells
        has_data = False
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            row_count += 1
            for cell in row:
                if cell is not None:
                    has_data = True
                    break
            if has_data:
                break
                
        if not has_data:
            print(f"  [ERROR] {filepath} is completely empty (no data cells found).")
            return False
            
        print(f"  [OK] {filepath} is valid (contains {len(wb.sheetnames)} sheet(s), active sheet has data).")
        return True
        
    except openpyxl.utils.exceptions.InvalidFileException as e:
        print(f"  [ERROR] {filepath} is not a valid Excel file or is corrupted: {e}")
        return False
    except Exception as e:
        print(f"  [ERROR] Unexpected error validating {filepath}: {e}")
        return False

def main():
    files_to_check = glob.glob("src/*/Методы оптимизации/*/source/*.xlsx")
    
    if not files_to_check:
        print("No Excel files found to validate.")
        sys.exit(0)
        
    print(f"Found {len(files_to_check)} Excel files to validate.\n")
    
    all_valid = True
    for file in files_to_check:
        if not validate_excel_file(file):
            all_valid = False
            
    if all_valid:
        print("\n✅ All Excel files passed validation successfully!")
        sys.exit(0)
    else:
        print("\n❌ One or more Excel files failed validation.")
        sys.exit(1)

if __name__ == "__main__":
    main()
