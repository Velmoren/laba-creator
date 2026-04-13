import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import os

def create_border():
    thin = Side(border_style="thin", color="FF000000")
    return Border(top=thin, left=thin, right=thin, bottom=thin)

def apply_corporate(ws, header_cells, input_cells, target_cell):
    header_fill = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True)
    input_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    target_fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
    
    for cell in header_cells:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].border = create_border()
        
    for cell in input_cells:
        ws[cell].fill = input_fill
        ws[cell].border = create_border()
        
    ws[target_cell].fill = target_fill
    ws[target_cell].border = create_border()

def apply_minimalist(ws, all_cells):
    bold_font = Font(bold=True)
    for cell in all_cells:
        ws[cell].border = create_border()
        if ws[cell].value and isinstance(ws[cell].value, str) and ("x" not in ws[cell].value.lower()):
            pass # Keep it simple

def apply_colorful(ws, header_cells, input_cells, target_cell, constraint_cells):
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True)
    input_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid") # Yellow
    target_fill = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid") # Green
    constraint_fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid") # Orange
    
    for cell in header_cells:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].border = create_border()
        
    for cell in input_cells:
        ws[cell].fill = input_fill
        ws[cell].border = create_border()
        
    ws[target_cell].fill = target_fill
    ws[target_cell].border = create_border()
    
    for cell in constraint_cells:
        ws[cell].fill = constraint_fill
        ws[cell].border = create_border()

# Student 2: Trushkin (Corporate, Vertical, Russian)
# LR1: Var 1.5
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Модель (Вариант 1.5)"
ws["B2"] = "Переменные (Га)"
ws["C2"] = "x1 (Кукуруза)"
ws["D2"] = "x2 (Соя)"
ws["B3"] = "Значения"
ws["C3"] = 0
ws["D3"] = 0

ws["B5"] = "Целевая функция"
ws["C5"] = "Прибыль (ден. ед.)"
ws["D5"] = "=SUMPRODUCT(C3:D3, C6:D6)"
ws["B6"] = "Коэффициенты"
ws["C6"] = 90
ws["D6"] = 360

ws["B8"] = "Ограничения"
ws["C8"] = "x1"
ws["D8"] = "x2"
ws["E8"] = "Левая часть"
ws["F8"] = "Знак"
ws["G8"] = "Правая часть (Запас)"

data_constraints = [
    ("Земля (га)", 1, 1, "<=", 400),
    ("Ссуда (ден. ед.)", 200, 100, "<=", 60000),
    ("Склад (ц)", 30, 60, "<=", 21000)
]

for i, (name, c1, c2, sign, rhs) in enumerate(data_constraints, start=9):
    ws[f"B{i}"] = name
    ws[f"C{i}"] = c1
    ws[f"D{i}"] = c2
    ws[f"E{i}"] = f"=SUMPRODUCT(C3:D3, C{i}:D{i})"
    ws[f"F{i}"] = sign
    ws[f"G{i}"] = rhs

apply_corporate(ws, ["B2", "C2", "D2", "B5", "C5", "D5", "B8", "C8", "D8", "E8", "F8", "G8"], ["C3", "D3"], "D5")
wb.save("src/2_trushkin_macbus/Методы оптимизации/ЛР_1/source/lab1_model.xlsx")

# LR2: Var 1.5
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Оптимизация (Вариант 1.5)"
ws["B2"] = "Переменные (т)"
ws["C2"] = "x1 (Молоко)"
ws["D2"] = "x2 (Кефир)"
ws["E2"] = "x3 (Сметана)"
ws["B3"] = "Значения"
ws["C3"] = 100
ws["D3"] = 0
ws["E3"] = 0

ws["B5"] = "Целевая функция"
ws["C5"] = "Прибыль (руб.)"
ws["D5"] = "=SUMPRODUCT(C3:E3,C6:E6)"
ws["B6"] = "Коэффициенты"
ws["C6"] = 30
ws["D6"] = 22
ws["E6"] = 136

ws["B8"] = "Ограничения"
ws["C8"] = "x1"
ws["D8"] = "x2"
ws["E8"] = "x3"
ws["F8"] = "Факт"
ws["G8"] = "Знак"
ws["H8"] = "Лимит"

constraints_2 = [
    ("Молоко-сырье (т)", 1.01, 1.01, 9.45, "<=", 136),
    ("Оборудование (ч)", 0.18, 0.19, 0, "<=", 21.4),
    ("Автоматы сметаны (ч)", 0, 0, 3.25, "<=", 16.25),
    ("План по молоку", 1, 0, 0, ">=", 100)
]
for i, (name, c1, c2, c3, sign, rhs) in enumerate(constraints_2, start=9):
    ws[f"B{i}"] = name
    ws[f"C{i}"] = c1
    ws[f"D{i}"] = c2
    ws[f"E{i}"] = c3
    ws[f"F{i}"] = f"=SUMPRODUCT(C3:E3, C{i}:E{i})"
    ws[f"G{i}"] = sign
    ws[f"H{i}"] = rhs

apply_corporate(ws, ["B2", "C2", "D2", "E2", "B5", "C5", "D5", "B8", "C8", "D8", "E8", "F8", "G8", "H8"], ["C3", "D3", "E3"], "D5")
wb.save("src/2_trushkin_macbus/Методы оптимизации/ЛР_2/source/lab2_model.xlsx")

# Student 3: Novoselz (Minimalist, Horizontal, English)
# LR1: Var 1.3
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Optimization Var 1.3"

ws["B2"] = "Variables"
ws["B3"] = "x (Details X)"
ws["B4"] = "y (Details Y)"
ws["C2"] = "Values"
ws["C3"] = 600
ws["C4"] = 0

ws["E2"] = "Objective"
ws["E3"] = "Revenue"
ws["E4"] = "=SUMPRODUCT(C3:C4,F3:F4)"
ws["F2"] = "Coeffs"
ws["F3"] = 30
ws["F4"] = 40

ws["H2"] = "Constraints"
ws["I2"] = "x"
ws["J2"] = "y"
ws["K2"] = "LHS"
ws["L2"] = "Sign"
ws["M2"] = "RHS"

cons_3_1 = [
    ("Labor (hrs)", 1, 2, "<=", 4000),
    ("Max X", 1, 0, "<=", 2250),
    ("Max Y", 0, 1, "<=", 1750),
    ("Rods (kg)", 2, 5, "<=", 10000),
    ("Sheet (kg)", 5, 2, "<=", 10000),
    ("Order X", 1, 0, ">=", 600),
    ("Union total", 1, 1, ">=", 1500)
]
for i, (name, cx, cy, sign, rhs) in enumerate(cons_3_1, start=3):
    ws[f"H{i}"] = name
    ws[f"I{i}"] = cx
    ws[f"J{i}"] = cy
    ws[f"K{i}"] = f"=I{i}*$C$3 + J{i}*$C$4"
    ws[f"L{i}"] = sign
    ws[f"M{i}"] = rhs

all_cells = ["B2", "B3", "B4", "C2", "C3", "C4", "E2", "E3", "E4", "F2", "F3", "F4", "H2", "I2", "J2", "K2", "L2", "M2"] + [f"{col}{r}" for r in range(3, 10) for col in ["H","I","J","K","L","M"]]
apply_minimalist(ws, all_cells)
wb.save("src/3_novoselz_drum/Методы оптимизации/ЛР_1/source/lab1_model.xlsx")

# LR2: Var 2.3 (Transport)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Transport Var 2.3"

ws["B2"] = "Cost Matrix"
for j in range(1, 6): ws[f"{chr(67+j)}2"] = f"B{j}"
ws["C3"] = 4; ws["D3"] = 2; ws["E3"] = 3; ws["F3"] = 4; ws["G3"] = 1
ws["C4"] = 2; ws["D4"] = 4; ws["E4"] = 3; ws["F4"] = 5; ws["G4"] = 6
ws["C5"] = 6; ws["D5"] = 5; ws["E5"] = 4; ws["F5"] = 6; ws["G5"] = 2
for i in range(1, 4): ws[f"B{i+2}"] = f"A{i}"

ws["B7"] = "Shipment Matrix (Variables)"
for j in range(1, 6): ws[f"{chr(67+j)}7"] = f"B{j}"
for i in range(1, 4): 
    ws[f"B{i+7}"] = f"A{i}"
    for j in range(1, 6):
        ws[f"{chr(67+j)}{i+7}"] = 0
    ws[f"I{i+7}"] = f"=SUM(D{i+7}:H{i+7})"

ws["I7"] = "Row Sum"
ws["J7"] = "Sign"
ws["K7"] = "Supply"
ws["J8"] = "="; ws["K8"] = 60
ws["J9"] = "="; ws["K9"] = 90
ws["J10"] = "="; ws["K10"] = 140

ws["B11"] = "Col Sum"
for j in range(1, 6): ws[f"{chr(67+j)}11"] = f"=SUM({chr(67+j)}8:{chr(67+j)}10)"
ws["B12"] = "Sign"
for j in range(1, 6): ws[f"{chr(67+j)}12"] = "="
ws["B13"] = "Demand"
ws["D13"] = 40; ws["E13"] = 30; ws["F13"] = 90; ws["G13"] = 80; ws["H13"] = 50

ws["J2"] = "Total Cost"
ws["J3"] = "=SUMPRODUCT(D3:H5, D8:H10)"

all_cells_tr = []
for r in range(2, 6):
    for c in range(2, 8): all_cells_tr.append(f"{chr(64+c)}{r}")
for r in range(7, 14):
    for c in range(2, 12): all_cells_tr.append(f"{chr(64+c)}{r}")
all_cells_tr.extend(["J2", "J3"])

apply_minimalist(ws, all_cells_tr)
wb.save("src/3_novoselz_drum/Методы оптимизации/ЛР_2/source/lab2_model.xlsx")

# Student 4: Veliev (Colorful, Vertical, Russian, Semantic)
# LR1: Var 1.2
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Краски (Вариант 1.2)"

ws["B2"] = "Переменные"
ws["C2"] = "Q_ext (Наружная)"
ws["D2"] = "Q_int (Внутренняя)"
ws["B3"] = "Производство (т)"
ws["C3"] = 0
ws["D3"] = 0

ws["B5"] = "Целевая функция"
ws["C5"] = "Доход (руб.)"
ws["D5"] = "=SUMPRODUCT(C3:D3, C6:D6)"
ws["B6"] = "Цены"
ws["C6"] = 3000
ws["D6"] = 2000

ws["B8"] = "Ограничения"
ws["C8"] = "Q_ext"
ws["D8"] = "Q_int"
ws["E8"] = "Факт"
ws["F8"] = "Знак"
ws["G8"] = "Лимит"

cons_4_1 = [
    ("Продукт А", 1, 2, "<=", 6),
    ("Продукт В", 2, 1, "<=", 8),
    ("Спрос разница", -1, 1, "<=", 1),
    ("Спрос внутр.", 0, 1, "<=", 2)
]
constraint_cells = []
for i, (name, c1, c2, sign, rhs) in enumerate(cons_4_1, start=9):
    ws[f"B{i}"] = name
    ws[f"C{i}"] = c1
    ws[f"D{i}"] = c2
    ws[f"E{i}"] = f"=SUMPRODUCT(C3:D3, C{i}:D{i})"
    ws[f"F{i}"] = sign
    ws[f"G{i}"] = rhs
    constraint_cells.extend([f"E{i}", f"F{i}", f"G{i}"])

apply_colorful(ws, ["B2", "C2", "D2", "B5", "C5", "D5", "B8", "C8", "D8", "E8", "F8", "G8"], ["C3", "D3"], "D5", constraint_cells)
wb.save("src/4_veliev_macbook/Методы оптимизации/ЛР_1/source/lab1_model.xlsx")

# LR2: Var 1.2 (Coal)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Уголь (Вариант 1.2)"

ws["B2"] = "Переменные (Доля)"
ws["C2"] = "P_a (Сорт А)"
ws["D2"] = "P_b (Сорт В)"
ws["E2"] = "P_c (Сорт С)"
ws["B3"] = "Значения"
ws["C3"] = 0
ws["D3"] = 0
ws["E3"] = 0

ws["B5"] = "Целевая функция"
ws["C5"] = "Цена смеси"
ws["D5"] = "=SUMPRODUCT(C3:E3,C6:E6)"
ws["B6"] = "Цена за т"
ws["C6"] = 30
ws["D6"] = 30
ws["E6"] = 45

ws["B8"] = "Ограничения"
ws["C8"] = "P_a"
ws["D8"] = "P_b"
ws["E8"] = "P_c"
ws["F8"] = "Факт"
ws["G8"] = "Знак"
ws["H8"] = "Лимит"

cons_4_2 = [
    ("Баланс долей", 1, 1, 1, "=", 1),
    ("Фосфор (%)", 0.06, 0.04, 0.02, "<=", 0.03),
    ("Зола (%)", 2.0, 4.0, 3.0, "<=", 3.25)
]
constraint_cells2 = []
for i, (name, c1, c2, c3, sign, rhs) in enumerate(cons_4_2, start=9):
    ws[f"B{i}"] = name
    ws[f"C{i}"] = c1
    ws[f"D{i}"] = c2
    ws[f"E{i}"] = c3
    ws[f"F{i}"] = f"=SUMPRODUCT(C3:E3, C{i}:E{i})"
    ws[f"G{i}"] = sign
    ws[f"H{i}"] = rhs
    constraint_cells2.extend([f"F{i}", f"G{i}", f"H{i}"])

apply_colorful(ws, ["B2", "C2", "D2", "E2", "B5", "C5", "D5", "B8", "C8", "D8", "E8", "F8", "G8", "H8"], ["C3", "D3", "E3"], "D5", constraint_cells2)
wb.save("src/4_veliev_macbook/Методы оптимизации/ЛР_2/source/lab2_model.xlsx")
